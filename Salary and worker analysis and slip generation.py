#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd

file_path = 'Salary-Calculation-Sheet-and-Salary-Slip-Template-in-Excel.xlsm'

try:
    xls = pd.ExcelFile(file_path)
    
    database_df = pd.read_excel(xls, sheet_name='Database')
    consolidation_df = pd.read_excel(xls, sheet_name='Consolidation')

    print("Database Sheet Head:")
    print(database_df.head())
    
    print("\nConsolidation Sheet Head:")
    print(consolidation_df.head())
except Exception as e:
    print(f"An error occurred: {e}")


# In[3]:


#Reshaping
database_df.columns = database_df.iloc[1]
database_df = database_df.drop([0, 1]).reset_index(drop=True)


print("Cleaned Database DataFrame:")
database_df = database_df.reset_index(drop=True)
database_df.head()


# In[4]:


distribution_stats_full = database_df.describe(include='all')
distribution_stats_full


# In[5]:


# Checking for duplicate records
duplicate_records = database_df.duplicated().sum()

# Checking for missing values
missing_values = database_df.isnull().sum()


# In[6]:


# printing number of duplicate records
duplicate_records = database_df.duplicated().sum()
print(f"Number of duplicate records: {duplicate_records}")


# In[7]:


# printing columns with missing values
missing_values = database_df.isnull().sum()
print("Missing values in each column:")
print(missing_values)


# In[7]:


database_df_cleaned = database_df.drop_duplicates(subset='Name')


# In[8]:


database_df_cleaned


# In[9]:


database_df_cleaned = database_df.dropna(subset=['Name'])


# In[10]:


database_df_cleaned


# In[11]:


consolidation_df = pd.read_excel('Salary-Calculation-Sheet-and-Salary-Slip-Template-in-Excel.xlsm', sheet_name='Consolidation')


# In[12]:


# Load the Consolidation sheet
consolidation_df = pd.read_excel(xls, sheet_name='Consolidation', header=2)

# Display the cleaned Consolidation DataFrame
print("Cleaned Consolidation DataFrame:")
print(consolidation_df)


# In[13]:


# 1. Total amount to be paid for the month
total_payable = database_df['NET PAYABLE'].sum()
print("Total Amount to be Paid :", total_payable)


# In[17]:


import pandas as pd
with pd.ExcelWriter('cleaned_database.xlsx') as writer:
    database_df_cleaned.to_excel(writer, sheet_name='database', index=False)
    
print("DataFrame exported successfully to 'cleaned_database.xlsx'.")


# In[2]:


import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

file_path = 'cleaned_database.xlsx'
database_df = pd.read_excel(file_path)

# Defining columns for calculation of gross salary and total deductions
gross_salary_columns = ['Basic Salary', 'DA', 'HRA', 'Conv Working', 'Medi Working', 'Special', 'Bonus', 'TA']
deduction_columns = ['Contribution to PF', 'Profession Tax', 'TDS', 'Salary Advance']

# Converting relevant columns to numeric datatype
for column in gross_salary_columns + deduction_columns:
    database_df[column] = pd.to_numeric(database_df[column], errors='coerce')

# Calculate Gross Salary, Total Deductions, and NET PAYABLE
database_df['Gross Salary'] = database_df[gross_salary_columns].sum(axis=1)
database_df['Total Deductions'] = database_df[deduction_columns].sum(axis=1)
database_df['NET PAYABLE'] = database_df['Gross Salary'] - database_df['Total Deductions']


# In[3]:


# 1. Total amount to be paid for the month
total_payable = database_df['NET PAYABLE'].sum()

# 2. Average salary for workers
average_salary = database_df['NET PAYABLE'].mean()

# 3. Total deductions for the month
total_deductions = database_df['Total Deductions'].sum()
deductions_distribution = database_df[deduction_columns].sum()

# 4. Total number of men and women
gender_distribution = database_df['Gender'].value_counts()

# 5. Total number of leaves taken that month
total_leaves_taken = database_df['Leave taken'].sum()

# 6. Gross salary breakdown
gross_salary_breakdown = database_df[gross_salary_columns].sum()

# 7. Highest and lowest salaries
highest_salary = database_df['NET PAYABLE'].max()
lowest_salary = database_df['NET PAYABLE'].min()

# 8. Distribution of designations
designation_distribution = database_df['Designation'].value_counts()

# Display results
results = {
    "Total Amount to be Paid": total_payable,
    "Average Salary": average_salary,
    "Total Deductions": total_deductions,
    "Deductions Distribution": deductions_distribution.to_dict(),
    "Gender Distribution": gender_distribution.to_dict(),
    "Total Leaves Taken": total_leaves_taken,
    "Gross Salary Breakdown": gross_salary_breakdown.to_dict(),
    "Highest Salary": highest_salary,
    "Lowest Salary": lowest_salary,
    "Designation Distribution": designation_distribution.to_dict()
}

for key, value in results.items():
    print(f"{key}: {value}\n")

#correlation matrix
correlation_columns = gross_salary_columns + deduction_columns + ['NET PAYABLE']
correlation_matrix = database_df[correlation_columns].corr()

plt.figure(figsize=(15, 10))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f")
plt.title('Correlation Matrix')
plt.show()


# In[4]:


# 9. Distribution of company names
company_distribution = database_df['Your Company Name'].value_counts()


# In[8]:


results={"Company Name Distribution": company_distribution.to_dict()}
for key, value in results.items():
    print(f"{key}: {value}\n")


# In[9]:


database_df['Gender_encoded'] = database_df['Gender'].astype('category').cat.codes
database_df['Designation_encoded'] = database_df['Designation'].astype('category').cat.codes


# In[14]:


# Box plot for Gender vs Net Payable
plt.subplot(1, 2, 1)
sns.boxplot(x='Gender', y='NET PAYABLE', data=database_df)
plt.title('Gender vs Net Payable')


# In[15]:


# Box plot for Gender vs Basic Salary
plt.subplot(1, 2, 2)
sns.boxplot(x='Gender', y='Basic Salary', data=database_df)
plt.title('Gender vs Basic Salary')


# In[16]:


# Box plot for Designation vs Net Payable
plt.figure(figsize=(12, 6))
plt.subplot(1, 2, 1)
sns.boxplot(x='Designation', y='NET PAYABLE', data=database_df)
plt.title('Designation vs Net Payable')
plt.xticks(rotation=90)


# In[20]:


#Box plot for Designation vs Basic Salary
plt.figure(figsize=(12, 6))
plt.subplot(1, 2, 2)
sns.boxplot(x='Designation', y='Basic Salary', data=database_df)
plt.title('Designation vs Basic Salary')
plt.xticks(rotation=90)


# In[22]:


#Clustered column for Gender vs Designation
plt.figure(figsize=(12, 6))
sns.countplot(x='Designation', hue='Gender', data=database_df)
plt.title('Gender vs Designation')
plt.xticks(rotation=90)
plt.show()


# In[34]:


import pandas as pd
from openpyxl import Workbook
import os

output_dir = "All_salary_slips"
os.makedirs(output_dir, exist_ok=True)

def generate_salary_slip(employee_data):
    wb = Workbook()
    ws = wb.active

    # Employee information
    ws['A1'] = "Company Name"
    ws['A2'] = "Salary Pay Slip For The Month Of " + employee_data['Month'] + " " + str(employee_data['Year'])

    ws['A4'] = "Name:"
    ws['B4'] = employee_data['Name']
    ws['A5'] = "Designation:"
    ws['B5'] = employee_data['Designation']

    # Basic Payment and Gross Salary Section
    ws['A7'] = "I"
    ws['B7'] = "Basic Payment"
    ws['C7'] = employee_data['Basic Salary']

    ws['A8'] = "II"
    ws['B8'] = "Dearness Allowance"
    ws['C8'] = employee_data['DA']

    ws['A9'] = "III"
    ws['B9'] = "House Rent Allowance"
    ws['C9'] = employee_data['HRA']

    ws['A10'] = "IV"
    ws['B10'] = "Conveyance"
    ws['C10'] = employee_data['Conv Working']

    ws['A11'] = "V"
    ws['B11'] = "Medical Expenses"
    ws['C11'] = employee_data['Medi Working']

    ws['A12'] = "VI"
    ws['B12'] = "Special"
    ws['C12'] = employee_data['Special']

    ws['A13'] = "VII"
    ws['B13'] = "Bonus"
    ws['C13'] = employee_data['Bonus']

    ws['A14'] = "VIII"
    ws['B14'] = "TA"
    ws['C14'] = employee_data['TA']

    ws['A15'] = "Total Gross Salary"
    ws['C15'] = employee_data['Gross Salary']

    # Deductions Section
    ws['A17'] = "Deductions"
    ws['A18'] = "a. Contribution to PF"
    ws['C18'] = employee_data['Contribution to PF']

    ws['A19'] = "b. Salary Advance"
    ws['C19'] = employee_data['Salary Advance']

    ws['A20'] = "c. Profession Tax"
    ws['C20'] = employee_data['Profession Tax']

    ws['A21'] = "d. TDS"
    ws['C21'] = employee_data['TDS']

    ws['A22'] = "Total Deductions"
    ws['C22'] = employee_data['Total Deductions']

    # Net Payable
    ws['A24'] = "NET PAYABLE"
    ws['C24'] = employee_data['NET PAYABLE']

    # Authorised by
    ws['A26'] = "Authorised Signatory:"
    ws['B26'] = employee_data['Name of Authorised Signatory']

    
    output_path = os.path.join(output_dir, f"{employee_data['Name']}_salary_slip.xlsx")
    wb.save(output_path)

# Creating salary slips for all employees
for index, row in database_df.iterrows():
    generate_salary_slip(row)

print("Salary slips generated successfully.")


# In[32]:


import pandas as pd
from openpyxl import Workbook
import os

output_dir = "salary_slips"
os.makedirs(output_dir, exist_ok=True)

def generate_salary_slip(employee_name):
    employee_data = database_df[database_df['Name'] == employee_name].iloc[0]
    
    wb = Workbook()
    ws = wb.active

    # employee info
    ws['A1'] = "Company Name"
    ws['A2'] = "Salary Pay Slip For The Month Of " + employee_data['Month'] + " " + str(employee_data['Year'])

    ws['A4'] = "Name:"
    ws['B4'] = employee_data['Name']
    ws['A5'] = "Designation:"
    ws['B5'] = employee_data['Designation']

    # Basic Payment and Gross Salary Section
    ws['A7'] = "I"
    ws['B7'] = "Basic Payment"
    ws['C7'] = employee_data['Basic Salary']

    ws['A8'] = "II"
    ws['B8'] = "Dearness Allowance"
    ws['C8'] = employee_data['DA']

    ws['A9'] = "III"
    ws['B9'] = "House Rent Allowance"
    ws['C9'] = employee_data['HRA']

    ws['A10'] = "IV"
    ws['B10'] = "Conveyance"
    ws['C10'] = employee_data['Conv Working']

    ws['A11'] = "V"
    ws['B11'] = "Medical Expenses"
    ws['C11'] = employee_data['Medi Working']

    ws['A12'] = "VI"
    ws['B12'] = "Special"
    ws['C12'] = employee_data['Special']

    ws['A13'] = "VII"
    ws['B13'] = "Bonus"
    ws['C13'] = employee_data['Bonus']

    ws['A14'] = "VIII"
    ws['B14'] = "TA"
    ws['C14'] = employee_data['TA']

    ws['A15'] = "Total Gross Salary"
    ws['C15'] = employee_data['Gross Salary']

    # Deductions Section
    ws['A17'] = "Deductions"
    ws['A18'] = "a. Contribution to PF"
    ws['C18'] = employee_data['Contribution to PF']

    ws['A19'] = "b. Salary Advance"
    ws['C19'] = employee_data['Salary Advance']

    ws['A20'] = "c. Profession Tax"
    ws['C20'] = employee_data['Profession Tax']

    ws['A21'] = "d. TDS"
    ws['C21'] = employee_data['TDS']

    ws['A22'] = "Total Deductions"
    ws['C22'] = employee_data['Total Deductions']

    # Net Payable
    ws['A24'] = "NET PAYABLE"
    ws['C24'] = employee_data['NET PAYABLE']

    # Authorised by
    ws['A26'] = "Authorised Signatory:"
    ws['B26'] = employee_data['Name of Authorised Signatory']

    
    output_path = os.path.join(output_dir, f"{employee_data['Name']}_salary_slip.xlsx")
    wb.save(output_path)


# In[33]:


# Example with one employee
employee_name = "Raj Sharma"  
generate_salary_slip(employee_name)

print(f"Salary slip for {employee_name} generated successfully.")

